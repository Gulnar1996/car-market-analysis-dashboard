import requests
import openpyxl
from openpyxl import Workbook
from time import sleep
import os

BASE_URL = "https://autonet.az/api/items/searchItem?page={}"

HEADERS = {
    "x-authorization": "00028c2ddcc1ca6c32bc919dca64c288bf32ff2a",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

FILENAME = "autonet_items.xlsx"


if os.path.exists(FILENAME):
    wb = openpyxl.load_workbook(FILENAME)
    ws = wb.active
    print(f"✅ Mövcud '{FILENAME}' faylına davam edilir...")
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Autonet Data"
    ws.append([
        "make", "model", "buraxilis_ili", "muherrikin_tipi", "muherrikin_hecmi",
        "ban_tipi", "at_gucu", "suret_qutusu", "oturuculuk", "yurus",
        "cityName", "price", "currency"
    ])


muherrik_tipi_map = {
    1: "Benzin", 2: "Dizel", 3: "Qaz", 4: "Elektro", 5: "Hibrid"
}
ban_tipi_map = {
    1: "Sedan", 2: "Hetcbek", 3: "Miniven", 4: "Universal", 5: "Ofrouder",
    6: "Limuzin", 7: "Kabriolet", 8: "Kupe", 9: "Roadster", 10: "Furqon",
    11: "Pikap", 12: "Avtobus", 13: "Dartqi", 14: "Mikroavtobus",
    15: "Motosiklet", 16: "Qolfkar", 17: "Van", 18: "Yük maşını"
}
suret_qutusu_map = {1: "Avtomat", 2: "Mexanika"}
oturuculuk_map = {1: "Ön", 2: "Arxa", 3: "Hamısı"}


response = requests.get(BASE_URL.format(1), headers=HEADERS)
if response.status_code != 200:
    print("❌ API-yə qoşulmaq mümkün olmadı!")
    exit()

data = response.json()
total_pages = data.get("last_page", 1)
print(f"📄 Toplam {total_pages} səhifə tapıldı.\n")


start_page = 1
if ws.max_row > 1:
    written_pages = (ws.max_row - 1) // 20
    start_page = written_pages + 1
    print(f"📘 Davam edilir: səhifə {start_page}-dən başlayır.")


for page in range(start_page, total_pages + 1):
    print(f"➡️  Səhifə {page} yüklənir...")
    response = requests.get(BASE_URL.format(page), headers=HEADERS)
    if response.status_code != 200:
        print(f"⚠️  Səhifə {page} alınmadı! Kod: {response.status_code}")
        continue

    json_data = response.json()
    for item in json_data.get("data", []):
        ws.append([
            item.get("make"),
            item.get("model"),
            item.get("buraxilis_ili"),
            muherrik_tipi_map.get(item.get("muherrikin_tipi"), "Bilinmir"),
            item.get("muherrikin_hecmi"),
            ban_tipi_map.get(item.get("ban_tipi"), "Bilinmir"),
            item.get("at_gucu"),
            suret_qutusu_map.get(item.get("suret_qutusu"), "Bilinmir"),
            oturuculuk_map.get(item.get("oturuculuk"), "Bilinmir"),
            item.get("yurus"),
            item.get("cityName"),
            item.get("price"),
            item.get("currency")
        ])

    
    wb.save(FILENAME)
    print(f"💾 Səhifə {page} uğurla '{FILENAME}' faylına əlavə olundu.\n")

    sleep(0.5)

print("✅ Bütün səhifələr uğurla Excel faylına yazıldı.")
